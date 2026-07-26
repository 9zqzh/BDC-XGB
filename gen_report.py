#!/usr/bin/env python3
"""Generate 50-epoch training analysis Excel report for BDC 2026 StockTransformer."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter

# ── All 50 epochs data (parsed from training output) ──
# Format: [epoch, train_loss, train_final_score, eval_loss, eval_final_score,
#          spearman_rho, topk_hit_rate, win_rate, final_score_std, max_daily_loss,
#          early_stop_raw, early_stop_smooth, early_stop_best, counter, lr, patience,
#          saved_best, speed_s_per_it]

data = [
    [1,  1.0551, -0.0112, 1.0417, 0.0527, 0.1000, 0.0222, 0.7222, 0.1360, -0.1233, 0.052714, 0.052714, 0.052714, 0, 9.50e-06, 15, True,  5.82],
    [2,  1.0496,  0.0159, 1.0343, 0.0126, 0.1116, 0.0000, 0.5000, 0.1188, -0.1735, 0.012601, 0.040680, 0.052714, 0, 9.34e-06, 15, False, 7.12],
    [3,  1.0489,  0.0212, 1.0300, 0.0141, 0.1176, 0.0000, 0.6111, 0.0837, -0.1933, 0.014080, 0.032700, 0.052714, 0, 9.18e-06, 15, False, 6.39],
    [4,  1.0441,  0.0318, 1.0271, 0.0252, 0.1222, 0.0000, 0.7222, 0.0686, -0.1065, 0.025158, 0.030437, 0.052714, 0, 9.02e-06, 15, False, 6.63],
    [5,  1.0426,  0.0204, 1.0232, 0.0372, 0.1267, 0.0000, 0.6667, 0.0931, -0.1036, 0.037211, 0.032469, 0.052714, 0, 8.86e-06, 15, False, 5.79],
    [6,  1.0412,  0.0217, 1.0233, 0.0264, 0.1235, 0.0000, 0.6667, 0.0819, -0.1036, 0.026408, 0.030651, 0.052714, 0, 8.70e-06, 15, False, 6.43],
    [7,  1.0401,  0.0417, 1.0232, 0.0418, 0.1249, 0.0111, 0.6667, 0.1107, -0.0711, 0.041796, 0.033994, 0.052714, 0, 8.54e-06, 15, False, 5.79],
    [8,  1.0378,  0.0371, 1.0212, 0.0144, 0.1226, 0.0111, 0.6111, 0.1235, -0.1295, 0.014448, 0.028131, 0.052714, 0, 8.38e-06, 15, False, 8.36],
    [9,  1.0353,  0.0289, 1.0219, 0.0328, 0.1182, 0.0000, 0.6111, 0.0781, -0.0741, 0.032813, 0.029535, 0.052714, 0, 8.22e-06, 15, False, 5.74],
    [10, 1.0354,  0.0269, 1.0226, 0.0087, 0.1145, 0.0000, 0.5000, 0.0769, -0.0884, 0.008683, 0.023280, 0.052714, 0, 8.06e-06, 15, False, 5.80],
    [11, 1.0318,  0.0325, 1.0230, -0.0138,0.1171, 0.0000, 0.2778, 0.0751, -0.1570, -0.013775,0.012163, 0.052714, 0, 7.90e-06, 15, False, 5.79],
    [12, 1.0324,  0.0329, 1.0219, -0.0108,0.1177, 0.0000, 0.4444, 0.0861, -0.1570, -0.010833,0.005264, 0.052714, 0, 7.74e-06, 15, False, 5.74],
    [13, 1.0324,  0.0427, 1.0215, -0.0021,0.1158, 0.0111, 0.3889, 0.1195, -0.1956, -0.002085,0.003059, 0.052714, 0, 7.58e-06, 15, False, 5.72],
    [14, 1.0309,  0.0351, 1.0203, 0.0066, 0.1197, 0.0111, 0.4444, 0.1238, -0.1492, 0.006621, 0.004128, 0.052714, 0, 7.42e-06, 15, False, 5.71],
    [15, 1.0298,  0.0333, 1.0207, -0.0027,0.1166, 0.0111, 0.5556, 0.1499, -0.2649, -0.002702,0.002079, 0.052714, 0, 7.26e-06, 15, False, 5.74],
    [16, 1.0283,  0.0295, 1.0207, 0.0336, 0.1120, 0.0667, 0.5000, 0.2778, -0.4919, 0.033581, 0.011530, 0.052714, 1, 7.44e-06, 18, False, 5.80],
    [17, 1.0257,  0.0516, 1.0201, 0.0159, 0.1171, 0.0111, 0.5556, 0.1263, -0.2174, 0.015894, 0.012839, 0.052714, 2, 7.28e-06, 18, False, 5.78],
    [18, 1.0255,  0.0382, 1.0199, -0.0078,0.1149, 0.0111, 0.5556, 0.1793, -0.3594, -0.007820,0.006641, 0.052714, 3, 7.12e-06, 18, False, 5.78],
    [19, 1.0232,  0.0481, 1.0199, 0.0145, 0.1125, 0.0556, 0.5556, 0.2467, -0.5604, 0.014456, 0.008986, 0.052714, 4, 6.96e-06, 19, False, 5.74],
    [20, 1.0242,  0.0452, 1.0182, 0.0019, 0.1187, 0.0111, 0.5000, 0.1582, -0.2378, 0.001912, 0.006864, 0.052714, 5, 6.80e-06, 19, False, 5.77],
    [21, 1.0216,  0.0375, 1.0204, 0.0229, 0.1122, 0.0556, 0.5000, 0.1953, -0.1703, 0.022858, 0.011662, 0.052714, 6, 6.64e-06, 19, False, 5.79],
    [22, 1.0205,  0.0524, 1.0183, 0.0391, 0.1188, 0.0444, 0.5556, 0.1527, -0.2583, 0.039135, 0.019904, 0.052714, 7, 6.48e-06, 20, False, 6.10],
    [23, 1.0222,  0.0416, 1.0168, 0.0646, 0.1187, 0.0667, 0.5000, 0.2068, -0.3135, 0.064585, 0.033308, 0.052714, 8, 6.32e-06, 20, True,  189.47],
    [24, 1.0197,  0.0479, 1.0164, 0.0655, 0.1190, 0.0778, 0.5000, 0.2245, -0.1976, 0.065501, 0.042966, 0.052714, 9, 6.16e-06, 21, True,  5.80],
    [25, 1.0195,  0.0512, 1.0175, 0.0439, 0.1153, 0.0778, 0.5556, 0.2699, -0.4168, 0.043915, 0.043251, 0.052714, 10,6.00e-06, 21, False, 5.77],
    [26, 1.0182,  0.0424, 1.0159, 0.0281, 0.1169, 0.0778, 0.5000, 0.2789, -0.4168, 0.028058, 0.038693, 0.052714, 11,5.84e-06, 22, False, 5.83],
    [27, 1.0175,  0.0538, 1.0172, 0.0631, 0.1155, 0.0889, 0.5000, 0.2630, -0.3568, 0.063095, 0.046014, 0.052714, 12,5.68e-06, 22, False, 5.81],
    [28, 1.0150,  0.0472, 1.0147, 0.0809, 0.1176, 0.0889, 0.5556, 0.2760, -0.3135, 0.080909, 0.056482, 0.056482, 0, 5.52e-06, 15, True,  5.79],
    [29, 1.0139,  0.0577, 1.0145, 0.0714, 0.1193, 0.0889, 0.6111, 0.2551, -0.3135, 0.071384, 0.060953, 0.060953, 0, 5.36e-06, 15, False, 5.82],
    [30, 1.0139,  0.0609, 1.0139, 0.0789, 0.1222, 0.0889, 0.5556, 0.2731, -0.3135, 0.078885, 0.066333, 0.066333, 0, 5.20e-06, 15, False, 5.80],
    [31, 1.0133,  0.0687, 1.0141, 0.0675, 0.1200, 0.0889, 0.5556, 0.2988, -0.4168, 0.067546, 0.066697, 0.066697, 0, 5.04e-06, 15, False, 5.78],
    [32, 1.0129,  0.0514, 1.0147, 0.0713, 0.1166, 0.0889, 0.6111, 0.2981, -0.4168, 0.071261, 0.068066, 0.068066, 0, 4.88e-06, 15, False, 5.74],
    [33, 1.0117,  0.0584, 1.0145, 0.0806, 0.1164, 0.0889, 0.5556, 0.2927, -0.3048, 0.080642, 0.071839, 0.071839, 0, 4.72e-06, 15, False, 5.73],
    [34, 1.0120,  0.0538, 1.0134, 0.1015, 0.1158, 0.0889, 0.6111, 0.3095, -0.3048, 0.101533, 0.080747, 0.080747, 0, 4.56e-06, 15, True,  5.73],
    [35, 1.0094,  0.0702, 1.0143, 0.0763, 0.1173, 0.0889, 0.6111, 0.2679, -0.3135, 0.076282, 0.079408, 0.080747, 1, 4.40e-06, 27, False, 5.78],
    [36, 1.0098,  0.0471, 1.0154, 0.1075, 0.1108, 0.0889, 0.5556, 0.3047, -0.3044, 0.107500, 0.087835, 0.087835, 0, 4.24e-06, 15, True,  5.74],
    [37, 1.0105,  0.0667, 1.0148, 0.0919, 0.1150, 0.0889, 0.6111, 0.2834, -0.3044, 0.091903, 0.089056, 0.089056, 0, 4.08e-06, 15, False, 96.72],
    [38, 1.0090,  0.0626, 1.0146, 0.0588, 0.1149, 0.0778, 0.5556, 0.3067, -0.4168, 0.058796, 0.079978, 0.089056, 1, 3.92e-06, 30, False, 8.43],
    [39, 1.0069,  0.0570, 1.0138, 0.0704, 0.1170, 0.0778, 0.6111, 0.2765, -0.3044, 0.070416, 0.077109, 0.089056, 2, 3.76e-06, 31, False, 6.91],
    [40, 1.0086,  0.0582, 1.0153, 0.0843, 0.1147, 0.0889, 0.6111, 0.2881, -0.3135, 0.084272, 0.079258, 0.089056, 3, 3.60e-06, 32, False, 6.94],
    [41, 1.0077,  0.0585, 1.0161, 0.0712, 0.1117, 0.0778, 0.6111, 0.2814, -0.3135, 0.071229, 0.076849, 0.089056, 4, 3.44e-06, 33, False, 6.81],
    [42, 1.0058,  0.0749, 1.0153, 0.0747, 0.1138, 0.0778, 0.6111, 0.2722, -0.3044, 0.074749, 0.076219, 0.089056, 5, 3.28e-06, 35, False, 5.78],
    [43, 1.0052,  0.0688, 1.0156, 0.0889, 0.1114, 0.0889, 0.6111, 0.2884, -0.3135, 0.088873, 0.080015, 0.089056, 6, 3.12e-06, 36, False, 5.79],
    [44, 1.0050,  0.0740, 1.0149, 0.0832, 0.1126, 0.0889, 0.6111, 0.2919, -0.3174, 0.083205, 0.080972, 0.089056, 7, 2.96e-06, 37, False, 5.74],
    [45, 1.0042,  0.0713, 1.0156, 0.0850, 0.1116, 0.0889, 0.6111, 0.2880, -0.3174, 0.085042, 0.082193, 0.089056, 8, 2.80e-06, 39, False, 5.76],
    [46, 1.0057,  0.0662, 1.0149, 0.0874, 0.1120, 0.0889, 0.6111, 0.2915, -0.3174, 0.087426, 0.083763, 0.089056, 9, 2.64e-06, 41, False, 5.77],
    [47, 1.0043,  0.0686, 1.0149, 0.0790, 0.1133, 0.0889, 0.5000, 0.2894, -0.3174, 0.079049, 0.082349, 0.089056, 10,2.48e-06, 43, False, 5.78],
    [48, 1.0061,  0.0543, 1.0157, 0.0871, 0.1101, 0.0889, 0.5556, 0.2939, -0.3174, 0.087129, 0.083783, 0.089056, 11,2.32e-06, 45, False, 6.15],
    [49, 1.0032,  0.0706, 1.0155, 0.0592, 0.1106, 0.0667, 0.5556, 0.2665, -0.3044, 0.059221, 0.076414, 0.089056, 12,2.16e-06, 47, False, 7.03],
    [50, 1.0032,  0.0715, 1.0164, 0.0795, 0.1094, 0.0778, 0.5556, 0.2821, -0.3174, 0.079510, 0.077343, 0.089056, 13,2.00e-06, 49, False, 7.00],
]

# ── Style constants ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
PHASE_FILLS = {
    "Phase 1: Rapid Learning (E1-E7)": PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid"),
    "Phase 2: Plateau (E8-E15)": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    "Phase 3: Slow Climb (E16-E22)": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
    "Phase 4: Breakthrough (E23-E36)": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "Phase 5: Stagnation (E37-E50)": PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
}
BEST_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
BEST_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SAVED_FILL = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def get_phase(epoch):
    if epoch <= 7: return "Phase 1: Rapid Learning (E1-E7)"
    if epoch <= 15: return "Phase 2: Plateau (E8-E15)"
    if epoch <= 22: return "Phase 3: Slow Climb (E16-E22)"
    if epoch <= 36: return "Phase 4: Breakthrough (E23-E36)"
    return "Phase 5: Stagnation (E37-E50)"

def apply_style(ws, row, col, value, font=None, fill=None, fmt=None, alignment=CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or NORMAL_FONT
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════
# Sheet 1: Full Metrics Table (50 Epochs)
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Full Metrics (50 Epochs)"

headers = [
    "Epoch", "Phase",
    "Train Loss", "Train Final Score",
    "Eval Loss", "Eval Final Score",
    "Spearman Rho", "TopK Hit Rate", "Win Rate",
    "FS Std Dev", "Max Daily Loss",
    "ES Raw", "ES Smooth", "ES Best",
    "Counter", "LR", "Eff Patience",
    "Saved Best?", "Speed (s/it)",
]

for c, h in enumerate(headers, 1):
    apply_style(ws1, 1, c, h, font=HEADER_FONT, fill=HEADER_FILL)

widths = [6, 32, 11, 13, 11, 13, 11, 12, 10, 11, 12, 11, 11, 11, 8, 12, 12, 12, 11]
for c, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

for r, row_data in enumerate(data, 2):
    epoch = row_data[0]
    phase = get_phase(epoch)
    phase_fill = PHASE_FILLS.get(phase)
    is_saved = row_data[17]

    values = [
        epoch, phase,
        row_data[1], row_data[2], row_data[3], row_data[4],
        row_data[5], row_data[6], row_data[7], row_data[8], row_data[9],
        row_data[10], row_data[11], row_data[12],
        row_data[13], row_data[14], row_data[15],
        "★ SAVED" if is_saved else "",
        row_data[16],
    ]
    formats = [
        "0", "", "0.0000", "0.0000", "0.0000", "0.0000",
        "0.0000", "0.0000", "0.0000", "0.0000", "0.0000",
        "0.000000", "0.000000", "0.000000",
        "0", "0.00E+00", "0",
        "", "0.00",
    ]

    for c, (v, fmt) in enumerate(zip(values, formats), 1):
        row_fill = None
        row_font = NORMAL_FONT
        if is_saved and c == 18:
            row_fill = SAVED_FILL
            row_font = BOLD_FONT
        elif phase_fill and c <= 18:
            row_fill = phase_fill
        if c == 6 and isinstance(v, (int, float)) and v >= 0.09:
            if not is_saved:
                row_font = BOLD_FONT
        apply_style(ws1, r, c, v, font=row_font, fill=row_fill, fmt=fmt)

ws1.freeze_panes = "C2"
ws1.conditional_formatting.add(
    f"F2:F51",
    DataBarRule(start_type="min", end_type="max", color="27AE60", showValue=True)
)

# ═══════════════════════════════════════════════════
# Sheet 2: Phase Analysis
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Phase Analysis")

apply_style(ws2, 1, 1, "50-Epoch Training Phase Analysis - StockTransformer (18-day Validation)",
            font=Font(name="Arial", size=14, bold=True, color="1F4E79"), alignment=LEFT)
ws2.merge_cells("A1:J1")
ws2.row_dimensions[1].height = 30

apply_style(ws2, 2, 1, "Model: StockTransformer | Task: Learning to Rank (CSI 300) | 341 train / 18 val samples | 2.04M params | LinearLR scheduler",
            font=Font(name="Arial", size=9, italic=True, color="7F8C8D"), alignment=LEFT)
ws2.merge_cells("A2:J2")

phase_headers = [
    "Phase", "Epochs", "Eval Loss Trend", "Spearman Trend",
    "Final Score Trend", "Std Dev Trend", "TopK Hit Rate",
    "Model State", "Key Insight"
]
for c, h in enumerate(phase_headers, 1):
    apply_style(ws2, 4, c, h, font=HEADER_FONT, fill=HEADER_FILL)

s2_widths = [30, 10, 16, 16, 16, 16, 14, 16, 50]
for c, w in enumerate(s2_widths, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

phase_data = [
    ["Phase 1: Rapid Learning (E1-E7)", "1-7",
     "1.042 to 1.023 down", "0.100 to 0.125 up",
     "0.053 to 0.042 (volatile)", "0.136 to 0.111",
     "0.000-0.022", "Healthy learning",
     "Model extracts core ranking signal. Eval loss drops every epoch. Spearman climbs 25%. Best epoch is E1 (final_score=0.053), not surpassed until E23."],
    ["Phase 2: Plateau (E8-E15)", "8-15",
     "1.021 to 1.021 flat", "0.123 to 0.117 down",
     "0.033 to -0.003 (near zero)", "0.078 to 0.150",
     "0.000-0.011", "Apparent stagnation",
     "ALL metrics flatline. final_score negative multiple times. Smoothed_score drops 0.053 to 0.002. NOT overfitting - Eval Loss is flat not rising. Model reorganizing parameters at lower LR."],
    ["Phase 3: Slow Climb (E16-E22)", "16-22",
     "1.021 to 1.018 slow down", "0.112 to 0.119 up",
     "0.034 to 0.039 (positive again)", "0.278 to 0.153",
     "0.011 to 0.067", "Recovery begins",
     "First awakening at E16: topk_hit_rate jumps to 0.067. Counter 1-7 but best_smoothed (0.053, E1) not yet beaten. LinearLR: 7.4e-6 to 6.5e-6. Fine-grained Top-5 discrimination learned."],
    ["Phase 4: Breakthrough (E23-E36)", "23-36",
     "1.017 to 1.015 down", "0.119 to 0.111 (flat/slight down)",
     "0.065 to 0.108 UP big", "0.207 to 0.305 UP",
     "0.067 to 0.089 (stable)", "Top-5 specialization",
     "E23 breaks E1 record. Best sequence: E23=0.065, E24=0.066, E28=0.081, E34=0.102, E36=0.108. Spearman peaks at E30 (0.122) then declines. Std dev DOUBLES. Model trades global ranking quality for top-5 precision - risky specialization."],
    ["Phase 5: Stagnation (E37-E50)", "37-50",
     "1.015 to 1.016 flat", "0.115 to 0.109 down",
     "0.092 to 0.060-0.088 (oscillating)", "0.283 to 0.282",
     "0.067 to 0.089", "Learning saturated",
     "E37 reaches best smoothed_score (0.089). After E38, all metrics plateau. Counter reaches 13 by E50. Spearman hits MINIMUM at E50 (0.109). LR decays to 2.0e-6. Continue training yields zero improvement. Early stopper would trigger at counter~15."],
]

for r, row_data in enumerate(phase_data, 5):
    ws2.row_dimensions[r].height = 72
    fill = PHASE_FILLS.get(row_data[0])
    for c, v in enumerate(row_data, 1):
        al = LEFT if c >= 8 else CENTER
        fl = fill if c <= 8 else None
        apply_style(ws2, r, c, v, font=NORMAL_FONT, fill=fl, alignment=al)

ws2.freeze_panes = "A5"

# ── Key Metrics at Checkpoints ──
row = 12
apply_style(ws2, row, 1, "Key Metrics at Checkpoint Epochs",
            font=Font(name="Arial", size=12, bold=True, color="1F4E79"), alignment=LEFT)
ws2.merge_cells(f"A{row}:H{row}")

row = 13
ck_headers = ["Epoch", "Eval Loss", "Final Score", "Spearman", "TopK Hit Rate",
              "FS Std Dev", "Max Daily Loss", "Win Rate"]
for c, h in enumerate(ck_headers, 1):
    apply_style(ws2, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)

checkpoint_epochs = [1, 5, 7, 10, 15, 16, 23, 28, 34, 36, 37, 43, 50]
for i, ep in enumerate(checkpoint_epochs):
    r = row + 1 + i
    d = data[ep - 1]
    vals = [ep, d[3], d[4], d[5], d[6], d[8], d[9], d[7]]
    fmts = ["0", "0.0000", "0.0000", "0.0000", "0.0000", "0.0000", "0.0000", "0.0000"]
    is_saved = d[17]
    for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
        f = BOLD_FONT if is_saved else NORMAL_FONT
        fl = SAVED_FILL if is_saved else None
        apply_style(ws2, r, c, v, font=f, fill=fl, fmt=fmt)

# ═══════════════════════════════════════════════════
# Sheet 3: Recommendations
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Recommendations")

apply_style(ws3, 1, 1, "BDC 2026 StockTransformer - Training Analysis and Recommendations",
            font=Font(name="Arial", size=14, bold=True, color="1F4E79"), alignment=LEFT)
ws3.merge_cells("A1:H1")

apply_style(ws3, 2, 1, "Generated from 50-epoch training run with 18-day sliding window validation",
            font=Font(name="Arial", size=9, italic=True, color="7F8C8D"), alignment=LEFT)
ws3.merge_cells("A2:H2")

rec_headers = ["#", "Category", "Finding", "Severity", "Recommendation", "Priority", "Expected Impact", "Notes"]
for c, h in enumerate(rec_headers, 1):
    apply_style(ws3, 4, c, h, font=HEADER_FONT, fill=HEADER_FILL)

s3_widths = [4, 16, 42, 10, 48, 10, 18, 32]
for c, w in enumerate(s3_widths, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

SEV_HIGH = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
SEV_MED  = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
SEV_LOW  = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
PRI_URG  = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
PRI_HIGH = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
PRI_MED  = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
PRI_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")

recommendations = [
    [1, "Model Selection",
     "Best epoch by final_score is E36 (0.108) but it has worst Spearman (0.111) and std=0.305. E28 (0.081, std=0.276) or E1 (0.053, std=0.136) may generalize better.",
     "HIGH",
     "DO NOT use E36 blindly. Run predict.py with all saved checkpoints against test set. Choose by TEST score, not VAL score.",
     "URGENT", "Decisive for competition result",
     "Current best_model.pth = E36. Manual comparison needed since only last best is saved."],

    [2, "Overfitting Risk",
     "Spearman declined from peak 0.127 (E5) to 0.109 (E50) while final_score doubled. Classic top-5 overfitting: sacrificing global ranking for noisy top-5 bets.",
     "HIGH",
     "Add composite score = 0.5*Spearman + 0.5*normalized_final_score for checkpoint selection. E1 would score higher on stability-adjusted metrics than E36.",
     "URGENT", "Prevents selecting overfitted checkpoint",
     "Requires ~5 lines change in train.py save logic."],

    [3, "Validation Stability",
     "final_score_std increased 2.2x from E1 (0.136) to E36 (0.305). max_daily_loss worsened from -0.12 to -0.30. Model varies wildly across days.",
     "HIGH",
     "Use stability-adjusted score = final_score / final_score_std. E1=0.39, E7=0.38, E36=0.35. E1 wins on risk-adjusted basis.",
     "HIGH", "More reliable model selection",
     "Standard quant finance practice."],

    [4, "Early Stopping Tuning",
     "15-epoch warmup too short (breakthrough at E23). Adaptive patience expanded to 49, disabling early stopper at late epochs.",
     "MEDIUM",
     "Increase warmup to 25. Cap effective_patience at 25. Reduce EMA alpha 0.3 to 0.2. Add Spearman as secondary criterion.",
     "MEDIUM", "Prevents wasted GPU hours",
     "E37-E50 = 11 percent of total time, zero net improvement."],

    [5, "Loss Function",
     "topk_hit_rate maxed at 0.089 - 4.4 out of 5 top picks missed. Pairwise top5_weight=2.0 too weak. ListMLE focuses on mid-ranking.",
     "MEDIUM",
     "Try: (1) Increase top5_weight from 2.0 to 5.0-10.0. (2) Add LambdaRank loss. (3) ListNet with larger weight on top positions.",
     "MEDIUM", "Could significantly improve top-5",
     "Competition only cares about your 5 picks."],

    [6, "Test Set Evaluation",
     "Model NOT yet evaluated on actual test set (2026-03-09 to 2026-03-13). All analysis is validation-set only.",
     "MEDIUM",
     "IMMEDIATE: Run predict.py then score_self.py with current best_model.pth. Record baseline. Then run cross_val.py for robustness.",
     "URGENT", "Ground truth for model quality",
     "Validation metrics can mislead. Test set is unbiased judge."],

    [7, "Training Efficiency",
     "E23 took 189 s/it (4.5 hours). E37 took 97 s/it. Likely CPU/GPU contention from other applications.",
     "LOW",
     "Close other apps during training. Reduce num_epochs to 35 (post-E36 is wasted). Monitor GPU memory during runs.",
     "LOW", "Faster iteration cycles",
     "4 outlier epochs consumed ~8 extra hours."],

    [8, "Label Engineering",
     "Current label = (open_t5 - open_t1) / open_t1. 5-day horizon may be too noisy for consistent top-stock ranking.",
     "LOW",
     "A/B test: try label = (close_t5 - close_t1) / close_t1 or multi-horizon ensemble. Compare via run_commands.md workflow.",
     "LOW", "Potential improvement path",
     "Only after establishing baseline test score."],
]

for i, rec in enumerate(recommendations):
    r = 5 + i
    ws3.row_dimensions[r].height = 85
    for c, v in enumerate(rec, 1):
        sev_map = {"HIGH": SEV_HIGH, "MEDIUM": SEV_MED, "LOW": SEV_LOW}
        pri_map = {"URGENT": PRI_URG, "HIGH": PRI_HIGH, "MEDIUM": PRI_MED, "LOW": None}
        sev_fill = sev_map.get(rec[3]) if c == 4 else None
        pri_fill = pri_map.get(rec[5]) if c == 6 else None
        pri_font = PRI_FONT if c == 6 and pri_fill else NORMAL_FONT
        if c in (3, 4, 7):  # text
            apply_style(ws3, r, c, v, font=NORMAL_FONT, alignment=LEFT)
        elif c == 4:
            apply_style(ws3, r, c, v, font=BOLD_FONT, fill=sev_fill)
        elif c == 6:
            apply_style(ws3, r, c, v, font=pri_font, fill=pri_fill)
        else:
            apply_style(ws3, r, c, v, font=NORMAL_FONT)

ws3.freeze_panes = "A5"

# Bottom line summary
sr = 5 + len(recommendations) + 2
apply_style(ws3, sr, 1, "BOTTOM LINE: What To Do Right Now",
            font=Font(name="Arial", size=12, bold=True, color="C0392B"), alignment=LEFT)
apply_style(ws3, sr+1, 1,
    "1. Run predict.py NOW with best_model.pth (E36) to get test score as baseline.\n"
    "2. Modify train.py to save per-epoch checkpoints, then retrain and compare E1/E5/E23/E28/E36 on test set.\n"
    "3. Switch checkpoint selection from pure final_score to stability-adjusted composite score.\n"
    "4. Increase pairwise top5_weight (2.0 to 5.0) to address topk_hit_rate ceiling at 0.089.\n"
    "5. Reduce num_epochs to 35 - E37-E50 contributed zero net improvement.",
    font=NORMAL_FONT, alignment=LEFT)
ws3.merge_cells(f"A{sr+1}:H{sr+1}")
ws3.row_dimensions[sr+1].height = 100

# Save
output_path = "/sessions/sleepy-confident-edison/mnt/BDC(XGB)/training_analysis_50epochs.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Output filename: training_analysis_50epochs.xlsx")
